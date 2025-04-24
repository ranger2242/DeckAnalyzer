import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

# Predefined background colors you want to match (only these)
COLOR_HEX_TO_NAME = {
    'd9ead3': 'green',
    'f9cb9c': 'orange',
    'fff2cc': 'yellow'
}

def get_fill_hex(cell):
    fill = cell.fill
    if fill is None or not isinstance(fill, PatternFill):
        return None
    fg = fill.fgColor
    if fg is None or fg.type != 'rgb' or fg.rgb is None:
        return None
    return fg.rgb[-6:].lower()  # e.g., 'd9ead3'

def extract_colored_rows(file_path):
    wb = load_workbook(file_path)
    os.makedirs("output", exist_ok=True)

    color_rows = defaultdict(list)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1):
            first_cell = row[0]
            hex_color = get_fill_hex(first_cell)
            if hex_color in COLOR_HEX_TO_NAME:
                name = str(first_cell.value).strip() if first_cell.value else None
                if name:
                    cname = COLOR_HEX_TO_NAME[hex_color]
                    color_rows[cname].append(name)

    for color, names in color_rows.items():
        names = sorted(set(names))
        with open(f"./output/{color}.txt", "w", encoding="utf-8") as f:
            for name in names:
                f.write(name + "\n")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python extract_by_color.py path_to_file.xlsx")
        sys.exit(1)

    extract_colored_rows(sys.argv[1])
    print("✅ Extracted color-based name lists to ./output/")
